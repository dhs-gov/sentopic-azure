import logging
import io
from io import BytesIO
from io import StringIO
import logging
import nltk
from globals import globalvars
from globals import globalutils
import json
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook


def clean(text):
    # Make sure text is not over maximum number of embedding tokens
    if len(nltk.word_tokenize(text)) > globalvars.MAX_TOKENS:
         # Get only the last sentence
        text = globalutils.get_last_sentence(text)
    text = text.replace('"', '')
    text = text.replace("'", '')
    text = text.replace('\r','')
    text = text.replace('\n','')
    text = text.replace('\t','')
    return text


# JSON: Don't remove quotes
def clean_json(text):
    # Make sure text is not over maximum number of embedding tokens
    if len(nltk.word_tokenize(text)) > globalvars.MAX_TOKENS:
         # Get only the last sentence
        text = globalutils.get_last_sentence(text)
    text = text.replace('\r','')
    text = text.replace('\n','')
    text = text.replace('\t','')
    return text


def get_json_payload(json_obj):
    try:
        docs = json_obj.get('documents')
        if not docs:
            return None, "[data_extractor.get_json_payload()] No 'documents' key found."
        data = []
        for doc in docs:
            text = doc.get('text')
            print("text: ", text)
            if not text:
                return None, "[data_extractor.get_json_payload()] No 'text' key found."
            data.append(clean_json(text))
            print("json data: ", data)
        if data:
            return data, None
        else:
            return None, "[data_extractor.get_json_payload()] No data found in JSON."
    except Exception as e:    
        return None, str(e)


def get_file_text(bytes, is_json):
    try:
        fh_bytes = io.BytesIO(bytes)
        fh = io.TextIOWrapper(fh_bytes, encoding='cp1252', errors='strict')
        reader = fh.readlines()
        data_text = ""
        for row in reader:
            text = "".join(row)
            cleaned = None
            if (is_json):
                cleaned = clean_json(text)
            else:
                cleaned = clean(text)
            data_text = data_text + cleaned
        return data_text, None
    except Exception as e:   
        return None, str(e)


def get_file_data(bytes, is_json):
    try:
        fh_bytes = io.BytesIO(bytes)
        fh = io.TextIOWrapper(fh_bytes, encoding='cp1252', errors='strict')
        reader = fh.readlines()
        data_list = []
        for row in reader:
            text = "".join(row)
            print("[data_util.convert_bytes()] text: ", text) 
            cleaned = None
            if (is_json):
                cleaned = clean_json(text)
            else:
                cleaned = clean(text)
            data_list.append(cleaned)
        return data_list, None
    except Exception as e:   
        return None, str(e)


# XLSX: Reads only one column
def get_xlsx_data(bytes):
    try:
        xlsx = io.BytesIO(bytes)
        wb = load_workbook(xlsx)
        #ws = wb['Sheet1']
        #wb = load_workbook(contents)
        print(wb.sheetnames)
        #ws = wb.get_sheet_by_name('Sheet1')
        ws = wb.worksheets[0]                
        data_list = []
        for col in ws.iter_cols(min_row=0, max_col=1, max_row=ws.max_row):
            for cell in col:
                val = cell.value
                if val:
                    cleaned = clean(val)
                    data_list.append(cleaned)
        if data_list:
            return data_list, None
        else:
            return None, "Could not extract XLSX data."
    except Exception as e:   
        print("[data_util.get_xlsx_data()] Exception: ", str(e)) 
        return None, str(e)


def get_data(req):
    try:
        # Check JSON (non-file) payload
        return get_json_payload(req.get_json())
    except Exception as e:    
        #Ignore exception and check for files
        pass
    try:
        # If no JSON payload, check files
        data_all = []
        for input_file in req.files.values():
            filename = input_file.filename
            print("Filename: ", filename)
            contents = input_file.stream.read()
            #print("byte contents: ", contents)
            data_list = []
            if filename == 'stopwords.txt':
                print("[data_extractor] Found stopwords.txt")
            elif filename.endswith('.txt'):
                data_list, error = get_file_data(contents, is_json=False)
            elif filename.endswith('.json'):
                data_text, error = get_file_text(contents, is_json=True)
                if data_text:
                    json_obj = json.loads(data_text)
                    data_list, error = get_json_payload(json_obj)
            elif filename.endswith('.csv'):
                data_list, error = get_file_data(contents, is_json=False)
            elif filename.endswith('.xlsx'):
                data_list, error = get_xlsx_data(contents)
 
            if data_list:
                # Use extend if merging one list into another
                data_all.extend(data_list)
            elif error:
                return None, error
        if data_all:
            print("data_all: ", data_all)
            return data_all, None
        else:
            return None, "[data_extractor] No JSON or file payload detected."
    except Exception as e:    
        print("[data_extractor] Exception: ", str(e))
        return None, str(e)



 